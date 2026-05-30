"""
HARTS Consulting — Automated Event Scout
Runs every Monday via GitHub Actions. Discovers, scores, and logs events.
"""

import logging
import re
import time
import json
from datetime import datetime, date
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from dateutil import parser as dateparser
from duckduckgo_search import DDGS
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from config import (
    EXCEL_FILE_PATH, LOG_FILE, TODAY, REMINDER_DAYS_BEFORE,
    RECIPIENT_EMAIL, SEARCH_QUERIES
)

# ─── Logging Setup ────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ─── Constants ────────────────────────────────────────────────────────────────
REQUEST_TIMEOUT = 15
REQUEST_DELAY = 1.5
MAX_RESULTS_PER_QUERY = 8
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# Score thresholds for colour formatting
SCORE_COLOURS = {
    9:  "1A7A2E",  # dark green
    8:  "70AD47",  # light green
    7:  "FFFF00",  # yellow
    6:  "FFFACD",  # pale yellow
}
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
COLUMNS = [
    "Event Name", "Start Date", "End Date", "Country", "City",
    "Format", "Organizer", "URL", "Description",
    "Relevance Score", "Notified", "Date Added",
]

# ─── Scoring keywords ─────────────────────────────────────────────────────────
HIGH_SCORE_KEYWORDS = [
    "gcc", "global capability center", "shared services", "global business services",
    "gbs", "org design", "operating model", "centre of excellence", "center of excellence",
    "coe", "post-m&a", "post m&a", "merger integration", "post acquisition",
    "board advisory", "board governance", "nasscom gcc", "sson", "everest group",
    "isg outsourcing", "cii industry", "enterprise transformation", "enterprise restructuring",
]
MID_SCORE_KEYWORDS = [
    "chro", "hr transformation", "executive coaching", "talent leadership",
    "business strategy", "shrm", "people matters", "icf coaching", "ethrworld",
    "workforce strategy", "talent acquisition", "rpo", "recruitment process outsourcing",
]
SKIP_KEYWORDS = [
    "cybersecurity", "devops", "software engineering", "ai/ml tools",
    "analytics platform", "data science", "machine learning conference",
    "hackathon", "student", "academic", "webinar series", "vendor showcase",
]

TRUSTED_ORGANIZERS = [
    "nasscom", "sson", "everest group", "isg", "deloitte", "ey ", "ernst & young",
    "cii", "shrm", "people matters", "icf", "ethrworld", "kpmg", "pwc",
]


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 1 — DISCOVER EVENTS
# ═══════════════════════════════════════════════════════════════════════════════

def discover_urls() -> list[str]:
    """Run all 50 search queries and return deduplicated event URLs."""
    log.info("PHASE 1 — Discovering URLs across %d queries …", len(SEARCH_QUERIES))
    seen: set[str] = set()
    urls: list[str] = []

    with DDGS() as ddgs:
        for i, query in enumerate(SEARCH_QUERIES, 1):
            log.info("  Query %02d/%02d: %s", i, len(SEARCH_QUERIES), query)
            try:
                results = ddgs.text(query, max_results=MAX_RESULTS_PER_QUERY)
                for r in results:
                    url = r.get("href", "")
                    if url and url not in seen and _is_valid_event_url(url):
                        seen.add(url)
                        urls.append(url)
            except Exception as exc:
                log.warning("  Search failed for '%s': %s", query, exc)
            time.sleep(REQUEST_DELAY)

    log.info("PHASE 1 complete — %d unique URLs collected.", len(urls))
    return urls


def _is_valid_event_url(url: str) -> bool:
    """Filter out obvious non-event URLs."""
    skip_domains = ["linkedin.com", "facebook.com", "twitter.com", "youtube.com",
                    "instagram.com", "wikipedia.org", "reddit.com"]
    try:
        domain = urlparse(url).netloc.lower()
        return not any(d in domain for d in skip_domains)
    except Exception:
        return False


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 2 — EXTRACT EVENT DETAILS
# ═══════════════════════════════════════════════════════════════════════════════

def extract_events(urls: list[str]) -> list[dict]:
    """Visit each URL and extract structured event data."""
    log.info("PHASE 2 — Extracting event details from %d pages …", len(urls))
    events: list[dict] = []

    for i, url in enumerate(urls, 1):
        log.info("  Scraping %03d/%03d: %s", i, len(urls), url[:80])
        try:
            resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            event = _parse_page(resp.text, url)
            if event:
                score = _score_event(event)
                if score >= 6:
                    event["Relevance Score"] = score
                    events.append(event)
                    log.info("    ✓ Kept (score %d): %s", score, event.get("Event Name", "?"))
                else:
                    log.info("    ✗ Skipped (score %d)", score)
        except requests.RequestException as exc:
            log.warning("  Fetch failed %s: %s", url, exc)
        except Exception as exc:
            log.warning("  Parse error %s: %s", url, exc)
        time.sleep(REQUEST_DELAY)

    log.info("PHASE 2 complete — %d events scored ≥6.", len(events))
    return events


def _parse_page(html: str, url: str) -> dict | None:
    """Extract event fields from HTML using JSON-LD → meta tags → body text."""
    soup = BeautifulSoup(html, "lxml")
    event = {"URL": url}

    # — Try JSON-LD structured data first —
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0]
            if isinstance(data, dict) and data.get("@type") in ("Event", "BusinessEvent"):
                event["Event Name"] = _clean(data.get("name", ""))
                event["Start Date"] = _fmt_date(data.get("startDate", ""))
                event["End Date"]   = _fmt_date(data.get("endDate", ""))
                loc = data.get("location", {})
                if isinstance(loc, dict):
                    event["City"]    = _clean(loc.get("address", {}).get("addressLocality", ""))
                    event["Country"] = _clean(loc.get("address", {}).get("addressCountry", ""))
                event["Organizer"] = _clean(
                    data.get("organizer", {}).get("name", "") if isinstance(data.get("organizer"), dict)
                    else str(data.get("organizer", ""))
                )
                event["Description"] = _clean(data.get("description", ""))[:500]
        except Exception:
            pass

    # — Fall back to meta / title tags —
    if not event.get("Event Name"):
        title = soup.find("meta", property="og:title") or soup.find("title")
        event["Event Name"] = _clean(
            title.get("content", "") if hasattr(title, "get") else (title.string or "")
        )

    if not event.get("Description"):
        desc = soup.find("meta", property="og:description") or soup.find("meta", attrs={"name": "description"})
        if desc:
            event["Description"] = _clean(desc.get("content", ""))[:500]

    # — Derive missing fields from body text —
    body_text = soup.get_text(separator=" ", strip=True)

    if not event.get("Start Date"):
        event["Start Date"] = _extract_date_from_text(body_text)

    if not event.get("Country"):
        event["Country"] = _extract_country(body_text)

    if not event.get("City"):
        event["City"] = _extract_city(body_text)

    event["Format"] = _extract_format(body_text)
    event.setdefault("Organizer", _extract_organizer(body_text, url))
    event.setdefault("End Date", "")
    event.setdefault("Description", body_text[:400] if not event.get("Description") else event["Description"])

    # — Reject pages without a name —
    if not event.get("Event Name") or len(event["Event Name"]) < 4:
        return None

    # — Skip past events —
    start = _parse_date_obj(event.get("Start Date", ""))
    if start and start.date() < TODAY:
        log.info("    ✗ Past event (%s)", event["Start Date"])
        return None

    return event


def _clean(text: str) -> str:
    return re.sub(r"\s+", " ", str(text)).strip()


def _fmt_date(raw: str) -> str:
    """Parse ISO or freeform date to DD/MM/YYYY."""
    try:
        dt = dateparser.parse(raw, ignoretz=True)
        return dt.strftime("%d/%m/%Y") if dt else ""
    except Exception:
        return ""


def _parse_date_obj(raw: str) -> datetime | None:
    try:
        return dateparser.parse(raw, dayfirst=True, ignoretz=True)
    except Exception:
        return None


def _extract_date_from_text(text: str) -> str:
    patterns = [
        r"\b(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})\b",
        r"\b(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*[\s,]+\d{4})\b",
        r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{1,2}[\s,]+\d{4}\b",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return _fmt_date(m.group(0))
    return ""


COUNTRY_HINTS = {
    "India": ["india", "bangalore", "bengaluru", "hyderabad", "mumbai", "delhi",
              "chennai", "pune", "gurugram"],
    "Singapore": ["singapore"],
    "UAE": ["dubai", "abu dhabi", "uae", "united arab emirates"],
    "UK": ["london", "united kingdom", "uk"],
    "USA": ["new york", "san francisco", "chicago", "atlanta", "united states", "usa"],
    "Germany": ["munich", "berlin", "germany"],
    "Australia": ["sydney", "melbourne", "australia"],
    "Malaysia": ["kuala lumpur", "malaysia"],
}


def _extract_country(text: str) -> str:
    t = text.lower()
    for country, hints in COUNTRY_HINTS.items():
        if any(h in t for h in hints):
            return country
    return ""


def _extract_city(text: str) -> str:
    cities = [
        "Bangalore", "Bengaluru", "Hyderabad", "Mumbai", "Delhi", "Chennai", "Pune",
        "Gurugram", "Singapore", "Dubai", "London", "New York", "San Francisco",
        "Sydney", "Kuala Lumpur", "Munich", "Berlin", "Chicago", "Atlanta",
    ]
    t = text.lower()
    for city in cities:
        if city.lower() in t:
            return city
    return ""


def _extract_format(text: str) -> str:
    t = text.lower()
    if "hybrid" in t:
        return "Hybrid"
    if any(w in t for w in ["virtual", "online", "webinar", "livestream"]):
        return "Online"
    return "In-Person"


def _extract_organizer(text: str, url: str) -> str:
    domain = urlparse(url).netloc.replace("www.", "")
    for org in TRUSTED_ORGANIZERS:
        if org in text.lower():
            return org.title()
    return domain.split(".")[0].title()


# ═══════════════════════════════════════════════════════════════════════════════
# SCORING
# ═══════════════════════════════════════════════════════════════════════════════

def _score_event(event: dict) -> int:
    """Return relevance score 1–10 based on content rules."""
    text = " ".join([
        event.get("Event Name", ""),
        event.get("Description", ""),
        event.get("Organizer", ""),
    ]).lower()

    # Immediate skip
    if any(kw in text for kw in SKIP_KEYWORDS):
        return 3

    # High score (9–10)
    if any(kw in text for kw in HIGH_SCORE_KEYWORDS):
        score = 9
        if any(org in text for org in TRUSTED_ORGANIZERS):
            score = 10
        return score

    # Mid score (6–8)
    if any(kw in text for kw in MID_SCORE_KEYWORDS):
        score = 7
        if any(org in text for org in TRUSTED_ORGANIZERS):
            score = 8
        return score

    return 4  # not relevant


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 3 — UPDATE EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def update_excel(events: list[dict]) -> int:
    """Load or create Excel, add new events, apply formatting. Returns count added."""
    log.info("PHASE 3 — Updating Excel: %s", EXCEL_FILE_PATH)

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
        existing_urls = {ws.cell(row=r, column=8).value for r in range(2, ws.max_row + 1)}
        log.info("  Loaded existing file (%d rows).", ws.max_row - 1)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Events"
        existing_urls = set()
        _write_header(ws)
        log.info("  Created new workbook.")

    added = 0
    for event in events:
        if event.get("URL") in existing_urls:
            log.info("  Duplicate skipped: %s", event.get("Event Name"))
            continue
        row = ws.max_row + 1
        _write_row(ws, row, event)
        _apply_row_colour(ws, row, event.get("Relevance Score", 6))
        existing_urls.add(event.get("URL"))
        added += 1

    _auto_fit_columns(ws)
    wb.save(EXCEL_FILE_PATH)
    log.info("PHASE 3 complete — %d new events added.", added)
    return added


def _write_header(ws):
    for col, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


def _write_row(ws, row: int, event: dict):
    values = [
        event.get("Event Name", ""),
        event.get("Start Date", ""),
        event.get("End Date", ""),
        event.get("Country", ""),
        event.get("City", ""),
        event.get("Format", ""),
        event.get("Organizer", ""),
        event.get("URL", ""),
        event.get("Description", ""),
        event.get("Relevance Score", ""),
        "",  # Notified (blank)
        TODAY.strftime("%d/%m/%Y"),
    ]
    for col, val in enumerate(values, 1):
        ws.cell(row=row, column=col, value=val)


def _apply_row_colour(ws, row: int, score: int):
    colour = SCORE_COLOURS.get(score, "FFFFFF")
    fill = PatternFill("solid", fgColor=colour)
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(row=row, column=col).fill = fill


def _auto_fit_columns(ws):
    col_widths = {1: 40, 2: 14, 3: 14, 4: 14, 5: 16, 6: 12,
                  7: 22, 8: 45, 9: 60, 10: 8, 11: 10, 12: 12}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 4 — REMINDERS (M365 — DISABLED)
# ═══════════════════════════════════════════════════════════════════════════════

def send_reminders():
    """
    TODO: Enable once Microsoft 365 credentials are configured.

    Required:
        CLIENT_ID     = "<your-app-client-id>"
        CLIENT_SECRET = "<your-app-client-secret>"
        TENANT_ID     = "<your-tenant-id>"

    Logic (when enabled):
        1. Load HARTS_Events.xlsx
        2. For each row where Notified is blank:
           - Parse Start Date
           - If Start Date is within REMINDER_DAYS_BEFORE days:
             - Send email via Graph API to RECIPIENT_EMAIL
             - Subject: "Event Reminder — HARTS | {Event Name} on {Start Date}"
             - Body: full event details
             - Set Notified = "YES" in the row
        3. Save updated workbook
    """
    # import msal
    # from msgraph.core import GraphClient
    # ... implementation goes here ...
    log.info("PHASE 4 — M365 reminders are currently disabled.")


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 5 — LOGGING SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def log_summary(new_events: int, errors: int):
    summary = (
        f"\n{'='*60}\n"
        f"  HARTS Event Scout — Run Summary\n"
        f"  Date       : {TODAY.strftime('%d %B %Y')}\n"
        f"  New events : {new_events}\n"
        f"  Errors     : {errors}\n"
        f"{'='*60}\n"
    )
    log.info(summary)
    print(summary)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    log.info("=" * 60)
    log.info("HARTS Event Scout starting — %s", TODAY.isoformat())
    log.info("=" * 60)

    errors = 0

    try:
        urls   = discover_urls()
    except Exception as exc:
        log.error("PHASE 1 failed: %s", exc)
        urls   = []
        errors += 1

    try:
        events = extract_events(urls)
    except Exception as exc:
        log.error("PHASE 2 failed: %s", exc)
        events = []
        errors += 1

    try:
        added  = update_excel(events)
    except Exception as exc:
        log.error("PHASE 3 failed: %s", exc)
        added  = 0
        errors += 1

    send_reminders()   # Phase 4 — currently a no-op

    log_summary(added, errors)


if __name__ == "__main__":
    main()
